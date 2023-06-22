from conlib import *
from openpyxl import Workbook, load_workbook
from tkinter import *
import time

wb = load_workbook("input/kopf.xlsx")
ws = wb.active

class EditValue:
    """To edit the value in a cell in a given excel file"""
    def run():
        """Opens the window with the GUI to edit the value in a given cell"""
        root = window(size=400,sz_lmt=10,title="Edit the Value")
        rht3 = Entry(root.obj,border=2,width=12)
        rht3.pack()
        def calc(lett:str="A",numb:int=1):
            global ws, wb
            return ws[f"{lett}{numb}"].value


        def update():
            """Update the shown feedback if saving the file was successful or why it failed"""
            try:
                obj = Label(root.obj,text=str(calc(SearchForValue.rht1.get()[0],SearchForValue.rht2.get()[0])))
            except:
                obj = Label(root.obj,text="Unvalid Index")
            obj.destroy()
            try:
                obj = Label(root.obj,text=str(calc(SearchForValue.rht1.get()[0],SearchForValue.rht2.get()[0])))
            except:
                obj = Label(root.obj,text="Unvalid Index")
            obj.pack()
            root.obj.update()

        button = Button(root.obj, text="Save", command=update,width=10)
        button.pack()

class SearchForValue:
    """Opens the window with the GUI to reveal the value of a cell in a given excel-file"""
    root = window(size=400,sz_lmt=10,title="Search for a Value")
    rht1 = Entry(root.obj,border=2,width=12)
    rht2 = Entry(root.obj,border=2,width=12)
    rht1.pack()
    rht2.pack()

    def calc(lett:str="A",numb:int=1):
        """Returns the value of a given cell"""
        global ws, wb
        return ws[f"{lett}{numb}"].value

    try:
        obj = Button(root.obj,text=str(calc(rht1.get()[0],rht2.get()[0])),command=EditValue.run)
    except:
        obj = Label(root.obj,text="Unvalid Index")

    def update():
        """Updates the showns value of the cell"""
        SearchForValue.obj.destroy()
        try:
            SearchForValue.obj = Button(SearchForValue.root.obj,text=str(SearchForValue.calc(SearchForValue.rht1.get()[0],SearchForValue.rht2.get()[0])),command=EditValue.run)
        except:
            SearchForValue.obj = Label(SearchForValue.root.obj,text="Unvalid Index")
        SearchForValue.obj.pack()
        SearchForValue.root.obj.update()

    button = Button(root.obj, text="Reveal Value", command=update,width=10)
    button.pack()

while True:
    SearchForValue.root.obj.update()